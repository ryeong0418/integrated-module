from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter


class ExcelAlignment:
    """
    ExcelAlignment class
    """

    ALIGN_VERTICAL = "align_vertical"
    ALIGN_WRAP_TEXT = "align_wrap_text"


class ExcelBorder:
    """
    ExcelBorder class
    """

    BORDER_NONE = None
    BORDER_DASHDOT = "dashDot"
    BORDER_DASHDOTDOT = "dashDotDot"
    BORDER_DASHED = "dashed"
    BORDER_DOTTED = "dotted"
    BORDER_DOUBLE = "double"
    BORDER_HAIR = "hair"
    BORDER_MEDIUM = "medium"
    BORDER_MEDIUMDASHDOT = "mediumDashDot"
    BORDER_MEDIUMDASHDOTDOT = "mediumDashDotDot"
    BORDER_MEDIUMDASHED = "mediumDashed"
    BORDER_SLANTDASHDOT = "slantDashDot"
    BORDER_THICK = "thick"
    BORDER_THIN = "thin"

    BLACK_RGB = "000000"


class ExcelWriter:
    """
    ExcelWriter class
    """

    def __init__(self):
        self.wb = None
        self.active_worksheet = None

    def create_workbook(self):
        """
        workbook 생성 및 default sheet를 active_worksheet 설정
        """
        self.wb = Workbook()
        self.active_worksheet = self.wb.active

    def set_active_sheet_name(self, sheet_name: str):
        """
        active_worksheet에 title 설정 함수 (20자리 제한)
        :param sheet_name: 시트 title 이름
        :return: (10자리 제한된) 시트 title 이름
        """
        if len(sheet_name) > 20:
            sheet_name = sheet_name[:20]

        self.active_worksheet.title = sheet_name
        return sheet_name

    def set_cell_width_columns(self, columns_width_dict: dict):
        """
        active_worksheet에 특정 컬럼의 width 설정 함수
        :param columns_width_dict: 컬럼 width dict
        """
        for column, width in columns_width_dict.items():
            self.active_worksheet.column_dimensions[column].width = width

    def autofit_column_size(self, columns=None, margin=2):
        """
        컬럼 자동 크기 조정 함수
        :param columns: column list
        :param margin: margin value
        """
        for i, column_cells in enumerate(self.active_worksheet.columns):
            is_ok = False
            if columns is None:
                is_ok = True
            elif isinstance(columns, list) and i in columns:
                is_ok = True

            if is_ok:
                length = max(len(str(cell.value)) for cell in column_cells)
                self.active_worksheet.column_dimensions[column_cells[0].column_letter].width = length + margin

    def set_value_from_pandas(
        self, df, start_row_index, align_vertical=False, contain_index=False, contain_header=False
    ):
        """
        active_worksheet에 데이터프레임의 값을 넣는 함수
        :param df: 출력 하려는 데이터프레임
        :param start_row_index: 시작 row index
        :param align_vertical: align vertical flag
        :param contain_index: index 포함 flag
        :param contain_header: header 포함 flag
        """
        for r in dataframe_to_rows(df, index=contain_index, header=contain_header):
            self.active_worksheet.append(r)

        if align_vertical:
            self._set_style_by_option(ExcelAlignment.ALIGN_VERTICAL, start_row_index, df)

    def set_style_by_option(self, df, start_row_index, style_option_dict: dict):
        """
        style 설정 함수
        :param df: style 설정 하려는 데이터 프레임
        :param start_row_index: 시작 row index
        :param style_option_dict: style 옵션 (ExcelAlignment class 사용)
        """
        for column_value, style in style_option_dict.items():
            self._set_style_by_option(ExcelAlignment.ALIGN_WRAP_TEXT, start_row_index, df, column_value)

    def set_border_by_option(self, df, start_row_index):
        """
        테두리 설정 함수
        :param df: 테두리 설정 하려는 데이터 프레임
        :param start_row_index: 시작 row index
        :return:
        """
        for cell in self.active_worksheet.iter_rows(min_row=start_row_index + 1, max_row=start_row_index + len(df) + 1):
            for cell_in_row in cell:
                cell_in_row.border = Border(
                    left=Side(border_style=ExcelBorder.BORDER_THIN, color=ExcelBorder.BLACK_RGB),
                    right=Side(border_style=ExcelBorder.BORDER_THIN, color=ExcelBorder.BLACK_RGB),
                    top=Side(border_style=ExcelBorder.BORDER_THIN, color=ExcelBorder.BLACK_RGB),
                    bottom=Side(border_style=ExcelBorder.BORDER_THIN, color=ExcelBorder.BLACK_RGB),
                )

    # 현재 사용하는 style은 alignment에 대해서만 적용
    def _set_style_by_option(self, alignment_style, start_row_index, df, column_to_style=None):
        """
        실제 style 설정 함수
        :param alignment_style: alignment style
        :param start_row_index: 시작 row index
        :param df: style 적용 데이터 프레임
        :param column_to_style: style 적용하려는 column
        """
        min_col = None
        max_col = None

        if alignment_style == ExcelAlignment.ALIGN_VERTICAL:
            alignment_style = Alignment(vertical="center")

        elif alignment_style == ExcelAlignment.ALIGN_WRAP_TEXT:
            alignment_style = Alignment(wrap_text=True)

            min_col = df.columns.get_loc(column_to_style) + 1
            max_col = df.columns.get_loc(column_to_style) + 1
        else:
            alignment_style = Alignment(vertical="top")

        for cell in self.active_worksheet.iter_rows(
            min_row=start_row_index + 1, max_row=start_row_index + len(df) + 1, min_col=min_col, max_col=max_col
        ):
            for cell_in_row in cell:
                cell_in_row.alignment = alignment_style

    def set_height_row(self, row, height):
        """
        active_worksheet row에 height 설정 함수
        :param row: row index
        :param height: height 설정값
        """
        self.active_worksheet.row_dimensions[row].height = height

    def set_freeze_panes(self, cell_value):
        """
        active_worksheet 셀 고정 설정 함수
        :param cell_value: cell value
        """
        self.active_worksheet.freeze_panes = cell_value

    def save_workbook(self, file_path, file_name):
        """
        workbook 저장 함수
        :param file_path: file path
        :param file_name: file name
        """
        self.wb.save(f"{file_path}/{file_name}")
        self.wb.close()

    def set_databar_by_value(self, df, start_row_index, target_columns):
        """
        Databar를 각 셀에 적용 하려는 함수
        :param df: Databar 적용 하려는 데이터 프레임
        :param start_row_index: 시작 row index
        :param target_columns: 적용하려는 타겟 컬럼
        """
        rule = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="5080C0")

        for target_column in target_columns:
            column_letter = get_column_letter(target_column)
            start_row = self.active_worksheet[column_letter][start_row_index + 1].coordinate
            end_row = self.active_worksheet[column_letter][start_row_index + len(df)].coordinate
            self.active_worksheet.conditional_formatting.add(f"{start_row}:{end_row}", rule)
            [self._set_cell_format(cell[0], "0.0%") for cell in self.active_worksheet[start_row:end_row]]

    def set_value_to_target(self, value, cell_idx, row=None, heigth=None):
        """
        특정 셀에 값을 넣는 함수
        :param value: 입력하려는 값
        :param cell_idx: 입력하려는 cell 인덱스
        :param row: style 적용을 위한 row indx (optional)
        :param heigth: 적용 하려는 height 값 (optional)
        """
        self.active_worksheet[cell_idx] = value

        # 해당 row에 style이 필요하면 style 추가하여 사용
        if row is not None:
            alignment_style = Alignment(wrap_text=True)

            self.active_worksheet[cell_idx].alignment = alignment_style
            self.active_worksheet.row_dimensions[row].height = heigth

    @staticmethod
    def _set_cell_format(cell, fmt):
        """
        cell format 적용 함수
        :param cell: 적용하려는 cell
        :param fmt: 적용하려는 숫자 포맷
        """
        cell.number_format = fmt
