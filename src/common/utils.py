import importlib.util
import argparse
import os
import pandas as pd
import time
import re
import cx_Oracle

from pathlib import Path
from datetime import datetime, timedelta
from sqlalchemy import URL
from src.common.constants import DbTypeConstants
from src.common.module_exception import ModuleException
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils.cell import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties


class SystemUtils:
    """
    SystemUtils class
    """

    @staticmethod
    def to_camel_case(snake_str):
        """
        모듈명 이름으로 클래스명을 획득하는 함수
        :param snake_str: 모듈명 snake 형식
        :return: 클래스명
        """
        components = snake_str.split("_")
        return "".join(x.title() for x in components[0:])

    @staticmethod
    def get_module_class(module_name, class_name, path):
        """
        모듈의 클래스를 가져오는 함수
        :param module_name: 모듈 이름
        :param class_name: 클래스 이름
        :param path: 모듈 경로
        :return: 해당 클래스
        """
        m = SystemUtils.get_module_from_file(module_name, path)
        c = getattr(m, class_name)
        return c

    @staticmethod
    def get_module_from_file(module_name, file_path):
        """
        모듈 파일 이름으로 모듈을 임포트 할 모듈 반환 함수
        :param module_name: 해당 모듈 파일 이름
        :param file_path: 해당 모듈 파일 경로
        :return: 해당하는 모듈
        """
        spec = importlib.util.spec_from_file_location(module_name, str(Path(file_path) / f"{module_name}.py"))
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    @staticmethod
    def get_start_args():
        """
        proc : 모듈 기능별 동작 수행을 위한 설정값
            (i : Initialize, e :  Extract, s : Summary, v : Visualization, b : batch) - required
        s_date : 분석 데이터 수집, 취합, 시각화 시작 날짜
        interval : 수행 날짜 기간
        모듈 기동 args 획독
        :return: args
        """
        parser = argparse.ArgumentParser(description="Smart Analyzer Start Args")

        parser.add_argument("--proc", required=True)
        parser.add_argument("--s_date")
        parser.add_argument("--interval")
        parser.add_argument("--s_time")
        parser.add_argument("--s_interval")

        args = parser.parse_args()
        return args

    @staticmethod
    def get_environment_variable():
        """
        배포 환경 별 config, logger 설정을 위한 환경 변수
        모듈의 특성상 local, dev는 환경 변수로 설정하고 현장에서 특별한 설정이 없을 때 prod로 사용할 수 있게 구성
        :return: env
        """
        return "prod" if os.environ.get("DEPLOY_ENV") is None else os.environ.get("DEPLOY_ENV")

    @staticmethod
    def byte_transform(b, to, bsize=1024):
        """
        byte를 kb, mb, gb, tb, eb로 변환하는 함수
        :param b: 변환하려는 byte 값
        :param to: 변환 하려는 단위 (k : kb, m : mb, g : gb, t : tb, e : eb)
        :param bsize: 변환 상수
        :return: 변환 하려는 단위의 반올림 값
        """
        a = {"k": 1, "m": 2, "g": 3, "t": 4, "p": 5, "e": 6}
        r = float(b)
        for i in range(a[to]):
            r = r / bsize
        return round(r, 2)

    @staticmethod
    def get_file_content_in_path(query_folder, sql_name):
        """
        파일 내용(sql) 추출 함수
        :param query_folder: 폴더명
        :param sql_name: 파일이름
        :return: sql 쿼리
        """
        with open(query_folder + "/" + sql_name, "r", encoding="utf-8") as file:
            sql_query = file.read()

        return sql_query

    @staticmethod
    def get_filenames_from_path(path: str, prefix: str = "", suffix: str = ""):
        """
        path에 모든 파일 이름을 가져오는 함수
        :param path: 파일 이름을 가져오려는 절대 경로
        :param prefix: 시작이 prefix를 포함하는 파일 (optional)
        :param suffix: 끝이 suffix를 포함하는 파일 (optional)
        :return: 파일 이름 list
        """
        if not os.path.exists(path):
            os.makedirs(path)

        if not os.path.isdir(path):
            return

        return [x for x in os.listdir(path) if str(x).startswith(prefix) and str(x).endswith(suffix)]

    @staticmethod
    def set_update_execute_log(result, start_tm, result_code, result_msg) -> dict:
        """
        ae_execute_log 저장 하기 위한 함수
        :param result: 기능 동작 결과
        :param start_tm: 기능 시작 시간
        :param result_code: 기능 동작 결과 코드
        :param result_msg: 기능 동작 결과 메세지
        :return: 결과 dict
        """
        result_dict = dict()
        result_dict["result"] = result
        result_dict["execute_end_dt"] = DateUtils.get_now_timestamp()
        result_dict["execute_elapsed_time"] = time.time() - start_tm
        result_dict["result_code"] = result_code
        result_dict["result_msg"] = result_msg

        return result_dict

    @staticmethod
    def get_file_export_args():
        """
        파일 추출 및 적재 기능 args
        proc :  no value - export, 'insert' - db insert
        :return: args
        """
        parser = argparse.ArgumentParser(description="File Export Args")

        parser.add_argument("--proc")
        args = parser.parse_args()
        return args

    @staticmethod
    def extract_tablename_in_filename(filename):
        """
        파일 이름에서 테이블 이름 추출 하는 함수
        :param filename: 파일 네임
        :return: 추출된 테이블 명
        """
        return filename.split(".")[0].split("-")[1]

    @staticmethod
    def apply_thin_border(ws, b_style):
        """
        :param ws: 엑셀 시트
        :return : thin_border 적용한 데이터 테이블
        """
        border_style = Border(
            left=Side(style=b_style), right=Side(style=b_style), top=Side(style=b_style), bottom=Side(style=b_style)
        )

        for row_rng in ws.rows:
            for cell in row_rng:
                if cell.value is not None:
                    cell.border = border_style

    @staticmethod
    def apply_column_width(ws, width_num):
        """
        excel column width 설정
        """
        for col in range(ws.min_column, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = width_num

    @staticmethod
    def get_folder_to_path(path):
        """
        path에서 folder만 추출 하는 함수
        :param path: path
        :return: folder 리스트
        """
        folder_list = []
        for folder in os.listdir(path):
            sub_folder = os.path.join(path, folder)

            if os.path.isdir(sub_folder):
                folder_list.append(folder)

        return folder_list


class TargetUtils:
    """
    TargetUtils class
    """

    @staticmethod
    def set_engine_param(target_config, extend_mode: bool = False):
        """
        분석 모듈 DB connection을 위한 SqlAlchemy engine 생성 설정 세팅 함수
        :param target_config: 각 타겟 config
        :param extend_mode: db 확장 분석 mode flag
        :return: sqlalchemy url object, conn_args
        """
        if extend_mode:
            collector_db_type = str(target_config["analysis_target_type"]).lower()
        else:
            collector_db_type = str(target_config["collector_db_type"]).lower()

        if collector_db_type == DbTypeConstants.POSTGRES:
            driver_name = "postgresql+psycopg2"
            conn_args = {
                "keepalives": 1,
                "keepalives_idle": 30,
                "keepalives_interval": 10,
                "keepalives_count": 5,
            }
        elif collector_db_type == DbTypeConstants.MSSQL:
            driver_name = "mssql+pymssql"
            conn_args = {}
        elif collector_db_type == DbTypeConstants.ORACLE:
            driver_name = "oracle+cx_oracle"

            if target_config["service_name"]:
                dsn = cx_Oracle.makedsn(
                    target_config["host"], target_config["port"], service_name=target_config["service_name"]
                )
            elif target_config["sid"]:
                dsn = cx_Oracle.makedsn(target_config["host"], target_config["port"], sid=target_config["sid"])
            else:
                ModuleException("E009")

            url_str = "{}://{}:{}@{}/?".format(driver_name, target_config["user"], target_config["password"], dsn)

            conn_args = {"encoding": "UTF-8", "nencoding": "UTF-8", "events": True}

            if str(target_config["role"]).lower() == "sysdba":
                conn_args["mode"] = cx_Oracle.SYSDBA

            return url_str, conn_args

        else:
            raise ModuleException("E008")

        url_object = URL.create(
            driver_name,
            username=target_config["user"],
            password=target_config["password"],
            host=target_config["host"],
            port=target_config["port"],
            database=target_config["sid"],
        )

        return url_object, conn_args


class InterMaxUtils:
    """
    InterMaxUtils class
    """

    @staticmethod
    def meta_table_value(table_name, df):
        """
        InterMax 데이터 메타 데이터 추출 시 특정 테이블에 기본값을 넣기 위한 함수
        :param table_name: 테이블 이름
        :param df: 추출된 데이터 데이터 프레임
        :return: 기본값을 넣은 데이터 프레임
        """
        if table_name == "ae_was_dev_map":
            df["isdev"] = 1

        return df


class MaxGaugeUtils:
    """
    MaxGaugeUtils class
    """

    @staticmethod
    def reconstruct_by_grouping(results, with_max_seq=False):
        """
        db sql text 재조합을 위한 함수
        :param results: seq가 동일한 데이터들의 전체 리스트
        :param with_max_seq: seq의 max값까지 추출하기 위한 flag
        :return: 재조합된 데이터프레임
        """
        agg_arg = {"sql_text": "".join}
        if with_max_seq:
            agg_arg["seq"] = "max"

        results_df = pd.DataFrame(results, columns=["sql_text", "partition_key", "sql_uid", "seq"])
        results_df = results_df.groupby(["sql_uid", "partition_key"], as_index=False).agg(agg_arg)
        results_df.drop(columns="partition_key", inplace=True)
        return results_df


class SqlUtils:
    """
    SqlUtils class
    """

    @staticmethod
    def remove_unnecess_char(df, target_c: str, des_c: str = None, contains_comma=False):
        """
        정규식을 이용한 /t, /n, /r, comma 치환 함수
        :param df: 원본 데이터프레임
        :param target_c: 대상 타겟 컬럼
        :param des_c: 목적지 컬럼 (optional) if None target_c
        :param contains_comma: comma 값 치환 여부
        :return: 치환된 데이터프레임
        """
        des_c = target_c if des_c is None else des_c

        comma_attr = [(",", " ")]

        repls = {r"\\t": " ", r"\\n": " ", r"\\r": " ", "\t": " ", "\n": " ", "\r": " "}

        if contains_comma:
            repls.update(comma_attr)

        rep = dict((re.escape(k), v) for k, v in repls.items())
        pattern = re.compile("|".join(rep.keys()))

        df[des_c] = df[target_c].apply(lambda x: pattern.sub(lambda m: rep[re.escape(m.group(0))], x))
        return df

    @staticmethod
    def rex_processing(df):
        """
        sql text를 정규식 처리를 위한 함수
        in 구문, values 구문, DateTime 형식, Date 형식
        :param df: 정규식 처리를 위한 DataFrame, 해당 컬럼은 sql_text
        :return: 정규식 변환 처리된 DataFrame
        """
        df["sql_text"] = df["sql_text"].str.replace(r"\s+in\s?\([^)]*\)", " in(:args:)", regex=True)
        df["sql_text"] = df["sql_text"].str.replace(r"\s+values\s?\([^)]*\)", " values(:args:)", regex=True)
        df["sql_text"] = df["sql_text"].str.replace(
            r"\d{2,4}\-\d{2}\-\d{2} \d{2}:\d{2}:\d{2}\.?\d{0,3}", ":DATETIME:", regex=True
        )
        df["sql_text"] = df["sql_text"].str.replace(
            r"\d{2,4}\.\d{2}\.\d{2} \d{2}:\d{2}:\d{2}\.?\d{0,3}", ":DATETIME:", regex=True
        )
        df["sql_text"] = df["sql_text"].str.replace(r"\d{4}\-\d{2}\-\d{2}", ":DATE:", regex=True)
        df["sql_text"] = df["sql_text"].str.replace(r"[\uAC00-\uD7A3]+", ":HAN:", regex=True)
        return df

    @staticmethod
    def sql_replace_to_dict(sql, replace_dict):
        """
        sql query에 동적 parameter를 set 하기 위한 함수
        :param sql: SQL 원본 쿼리
        :param replace_dict: 변경이 필요한 파라미터의 dict
                (ex) table_name를 치환하기 위해서는 원본 쿼리에 #(table_name) 형식으로 만들어 준다
        :return: 치환된 sql query
        :exception: 치환되지 않는 형태의 오류 발생시 원본 쿼리와 각 동적 parameter dict 키 매핑 확인
        """
        for key in replace_dict.keys():
            sql = sql.replace(f"#({key})", replace_dict[key])

        return sql

    @staticmethod
    def get_sql_text_in_file(sql_file):
        """
        tuning sql text file를 로드하여 sql text를 만드는 함수
        :param sql_file: tuning sql text file
        :return: sql text
        """
        sql = []
        with open(sql_file, "r", encoding="utf-8") as file:
            while True:
                line = file.readline()
                if not line:
                    break

                sql.append(line.strip())
            file.close()

        return " ".join(sql)


class DateUtils:
    """
    DateUtils class
    """

    @staticmethod
    def get_date_by_interval(interval, fmt="%Y%m%d"):
        """
        interval에 따른 날짜를 구하기 위한 함수
        :param interval: 필요한 날짜 interval (ex, 어제 -1, 내일 1)
        :param fmt: return 포맷 (기본 : yyyymmdd)
        :return: interval 날짜
        """
        now = datetime.now()
        return (now + timedelta(days=interval)).strftime(fmt)

    @staticmethod
    def get_now_timestamp(fmt="%Y%m%d%H%M%S"):
        """
        현재 시간 timestamp 함수
        :param fmt: 포맷 optional (기본 : %Y%m%d%H%M%S)
        :return: 현재 시간 str (기본 14 자리)
        """
        return datetime.now().strftime(fmt)

    @staticmethod
    def get_each_date_by_interval(s_date, interval, arg_fmt="%Y%m%d"):
        """
        input_date에서 interval 이후 날짜를 구하기 위한 함수
        :param s_date: 시작 날짜
        :param interval: 시간 간격
        :param arg_fmt: 시작 날짜 format
        :return: 시작날짜 , 끝날짜
        """
        s_date = datetime.strptime(str(s_date), arg_fmt)
        e_date = s_date + timedelta(days=int(interval))
        return s_date, e_date

    @staticmethod
    def set_date_conditions_by_interval(input_date, input_interval, return_fmt):
        """
        input_date에서 interval 사이에 date값을 구하기 위한 함수
        :param input_date: 입력 받은 시작 date
        :param input_interval: 시간 간격
        :param return_fmt: return date format
        :return: date range list
        """
        date_conditions = []

        for i in range(0, int(input_interval)):
            from_date = datetime.strptime(str(input_date), "%Y%m%d")
            date_condition = from_date + timedelta(days=i)
            date_condition = date_condition.strftime(return_fmt)
            date_conditions.append(date_condition)

        return date_conditions

    @staticmethod
    def get_each_date_by_interval2(s_date, interval, s_time, time_interval, arg_fmt):
        """
        input_date에서 interval 이후 날짜를 구하기 위한 함수
        :param s_date: 시작 날짜
        :param interval: 시간 간격
        :param arg_fmt: 시작 날짜 format
        :return: 시작날짜 , 끝날짜
        """
        s_date = datetime.strptime(str(s_date), "%Y%m%d")  # 2023-11-26 00:00:00
        e_date = s_date + timedelta(days=int(interval) - 1)
        s_date = s_date.strftime(arg_fmt)  # 2023-11-26
        e_date = e_date.strftime(arg_fmt)
        e_time = str(int(s_time) + int(time_interval))

        return s_date, e_date, s_time, e_time


class ExcelUtils:
    """
    ExcelUtils class
    """

    @staticmethod
    def excel_export(excel_file, sheet_name_txt, df):
        """
        visualization 엑셀 export 함수
        :param excel_file: 엑셀 파일명
        :param sheet_name_txt: 엑셀 sheet 이름
        :param df: 추출하려는 데이터 데이터 프레임
        """
        if not os.path.exists(excel_file):
            with pd.ExcelWriter(excel_file, mode="w", engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name_txt, index=False)
        else:
            with pd.ExcelWriter(excel_file, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
                df.to_excel(writer, sheet_name=sheet_name_txt, index=False)

    @staticmethod
    def append_df_into_excel(excel_file_path, sheet_name, df, s_col, s_row, sheet_append_mode):
        """
        pandas를 이용해서 data를 excel에 insert함
        """
        writer = pd.ExcelWriter(excel_file_path, mode="a", engine="openpyxl", if_sheet_exists=sheet_append_mode)
        df.to_excel(writer, sheet_name=sheet_name, index=False, startcol=s_col, startrow=s_row)
        writer.save()
        writer.close()

    @staticmethod
    def create_excel_and_sheet(excel_file_path, metric_name_list):
        """
        excel worksheet 생성
        """
        wb = Workbook()

        for idx, i in enumerate(metric_name_list):
            wb.create_sheet(i, idx)

        wb.remove(wb["Sheet"])
        wb.save(excel_file_path)
        wb.close()

    @staticmethod
    def set_linechart_object(metric_name):
        """
        Linechart생성 및 그래프 설정 함수
        """
        line_chart = LineChart()
        line_chart.style = "10"
        line_chart.width = 40
        line_chart.height = 8
        line_chart.title = metric_name

        y_axis_color = GraphicalProperties(ln=LineProperties(solidFill="FFFFFF"))
        line_chart.y_axis.spPr = y_axis_color

        font_test = Font(typeface="Calibri")
        cp = CharacterProperties(latin=font_test, sz=600)
        line_chart.x_axis.textProperties = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
        line_chart.y_axis.textProperties = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
        line_chart.legend.textProperties = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

        title_layout = ManualLayout(xMode="edge", yMode="edge", x=0.02, y=0.02)
        line_chart.title.layout = Layout(manualLayout=title_layout)

        gridlines_color = GraphicalProperties(ln=LineProperties(solidFill="BFBFBF"))
        line_chart.y_axis.majorGridlines.spPr = gridlines_color

        return line_chart

    @staticmethod
    def set_data_and_category(ws, category, col, line_chart_type):
        """
        linechart에 데이터 및 카테고리 insert
        """
        for idx, cell_col in enumerate(col):
            data = Reference(
                ws, min_col=cell_col.column, max_col=cell_col.column, min_row=ws.min_row, max_row=ws.max_row
            )
            line_chart_type.add_data(data, titles_from_data=True)
            line_chart_type.set_categories(category)

    @staticmethod
    def set_series_marker_style(line_chart_series):
        """
        차트 series별로 marker스타일 및 글자크기 지정
        """
        colors = ["3F526C", "C61D51"]

        for indx, series in enumerate(line_chart_series):
            series.marker.symbol = "circle"
            series.marker.graphicalProperties.line.solidFill = "FFFFFF"
            series.graphicalProperties.line.width = 28553
            legend_name = f"INSTANCE-{indx + 1}"
            series.tx.strRef.f = f'"{legend_name}"'
            font_test = Font(typeface="Calibri")

            series.dLbls = DataLabelList()
            series.dLbls.showVal = True
            series.dLbls.dLblPos = "t"

            if indx < len(colors):
                series.marker.graphicalProperties.solidFill = colors[indx]
                series.graphicalProperties.line.solidFill = colors[indx]

                cp1 = CharacterProperties(latin=font_test, sz=800, b=True, solidFill=colors[indx])
                series.dLbls.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp1), endParaRPr=cp1)])

            else:
                series.marker.graphicalProperties.solidFill = "FFFF00"
