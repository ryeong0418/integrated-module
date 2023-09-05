from src import common_module as cm
from src.common.constants import SystemConstants
from src.common.utils import DateUtils, SystemUtils, SqlUtils, ExcelUtils

import os
import re
import time
from src.analysis_extend_target import OracleTarget
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter


class MetricPerformanceReport(cm.CommonModule):

    """
    AE_DB_SYSMETRIC 테이블에 있는 데이터 가져와서
    EXCEL에 데이터 및 그래프 INSERT
    """

    def __init__(self, logger):
        super().__init__(logger=logger)
        self.ot: OracleTarget = None

    def main_process(self):

        self.logger.debug("metric performance report")
        self._init_sa_target()
        self._insert_datatable_or_chartgraph()

    def _convert_sql_to_df(self, sql_path, filename):

        """
        sql query문을 dataframe 형태로 변환
        """

        s_date, e_date = DateUtils.get_each_date_by_interval2(
            self.config["args"]["s_date"], self.config["args"]["interval"], arg_fmt="%Y-%m-%d"
        )
        date_dict = {"StartDate": s_date, "EndDate": e_date}
        query = SystemUtils.get_file_content_in_path(sql_path, filename+".txt")
        date_query = SqlUtils.sql_replace_to_dict(query, date_dict)

        for df in self.st.get_data_by_query(date_query):
            df.columns = [i.upper() for i in df.columns]
            return df

    def _insert_datatable_or_chartgraph(self):

        """
        column txt 파일 읽어서 column명 추출
        {number} txt 파일명 추출
        """

        sql_path = f"{self.config['home']}/" + SystemConstants.CHART_SQL
        excel_path = f"{self.config['home']}/" + SystemConstants.CHART_EXCEL
        txt_file_list = SystemUtils.get_filenames_from_path(sql_path)

        for file in txt_file_list:
            filename = file.split(".")[0] #AE_DB_SQLSTAT-1
            excel_file_path = f"{excel_path}/{filename}.xlsx"

            if re.search(r'-(\d+)', filename):
                df = self._convert_sql_to_df(sql_path, filename)
                ExcelUtils.excel_export(excel_file_path, 'Sheet1', df)
                self._apply_excel_style(['Sheet1'], excel_file_path)

            if re.search(r'CHART',filename):
                table_name = filename.split("-")[0]
                column_filename = f"{table_name}-COLUMN.txt" # AE_DB_SYSMETRIC-COLUMN.txt / AE_DB_TIME_MODEL-COLUMN.txt

                query = SystemUtils.get_file_content_in_path(sql_path, column_filename)
                re_query = re.sub("\n", "", query)  # python에서 엔터는 \n으로 기입된다. \n 제거해주는 코드
                sheet_name_list = re_query.split(",")

                self.check_excel_format(sheet_name_list, excel_file_path, sql_path, filename)

    def check_excel_format(self,sheet_name_list, excel_file_path, sql_path, filename):

        """
        excel 파일이 있으면 해당 파일에 데이터 overwrite하고
        excel 파일이 없으면 새로 생성하여 데이터 insert 한다.
        """

        if os.path.exists(excel_file_path):
            self._check_sheet_name_list(sheet_name_list, excel_file_path, sql_path, filename)
            self._insert_linechart_from_data(sheet_name_list, excel_file_path)
            self._apply_excel_style(sheet_name_list, excel_file_path)

        else:
            self._make_excel_sheet_data(sheet_name_list, excel_file_path, sql_path, filename)
            self._insert_linechart_from_data(sheet_name_list, excel_file_path)
            self._apply_excel_style(sheet_name_list, excel_file_path)

    def _make_excel_sheet_data(self, sheet_name_list, excel_file_path, sql_path, filename):

        """
        df group by instance
        """

        if len(sheet_name_list) != 0:
            df = self._convert_sql_to_df(sql_path, filename)
            unique_instance_numbers = df['INSTANCE_NUMBER'].unique()

            for idx, instance_num in enumerate(unique_instance_numbers):
                instance_df = df[df['INSTANCE_NUMBER'] == instance_num]

                for sheet_name in sheet_name_list:
                    col_start = SystemUtils.arithmetic_sequence(1, 4, idx+1)
                    extract_df = instance_df[['DATE_TIME', 'INSTANCE_NUMBER', sheet_name]]
                    ExcelUtils.excel_export_append_overlay(excel_file_path, sheet_name, extract_df, col_start, 16)

    def _check_sheet_name_list(self, sheet_name_list, excel_file_path, sql_path, filename):

        """
        sheet_name check
        """

        wb = load_workbook(excel_file_path, data_only=True)

        not_exist_sheet_list = []
        exist_sheet_list = []

        for sheet_name in sheet_name_list:

            if sheet_name in wb.get_sheet_names():
                exist_sheet_list.append(sheet_name)
            else:
                not_exist_sheet_list.append(sheet_name)

        self._make_excel_sheet_data(not_exist_sheet_list, excel_file_path, sql_path, filename)
        self._overwrite_excel_sheet(exist_sheet_list, excel_file_path, sql_path, filename)

    def _overwrite_excel_sheet(self, sheet_name_list, excel_file_path, sql_path, filename):

        """
        기존에 있는 excel을 읽어서
        데이터 및 차트 업데이트
        """

        wb = load_workbook(excel_file_path, data_only=True)

        df = self._convert_sql_to_df(sql_path, filename)
        unique_instance_numbers = df['INSTANCE_NUMBER'].unique()
        for sheet_name in sheet_name_list:
            ws = wb[sheet_name]

            for idx, instance_num in enumerate(unique_instance_numbers):

                instance_df = df[df['INSTANCE_NUMBER'] == instance_num]
                extract_df = instance_df[['DATE_TIME', 'INSTANCE_NUMBER', sheet_name]].head()
                col = [cell for cell in ws[ws.min_row] if cell.value == "DATE_TIME"]

                col_indx = None
                row_indx = None
                try:
                    col_indx = col[idx].column
                    row_indx = col[idx].row

                except IndexError:
                    col_indx = SystemUtils.arithmetic_sequence(1, 5, idx + 1)

                finally:
                    if idx == 0:
                        ExcelUtils.insert_df_into_excel(excel_file_path, sheet_name, extract_df, col_indx-1, row_indx-1,
                                                        "a", "replace")

                    else:
                        ExcelUtils.insert_df_into_excel(excel_file_path, sheet_name, extract_df, col_indx-1, row_indx-1,
                                                        "a", "overlay")

    def _apply_excel_style(self, sheet_name_list, excel_file_path):

        """
        excel에 dataframe 기입시 스타일 지정
        table border_style, column width 지정
        """

        wb = load_workbook(excel_file_path, data_only=True)

        for sheet_name in sheet_name_list:
            ws = wb[sheet_name]

            SystemUtils.apply_thin_border(ws, wb, excel_file_path, "thin")
            time.sleep(1)
            SystemUtils.apply_column_width(ws, wb, excel_file_path, 20)

    def _insert_linechart_from_data(self, sheet_name_list, excel_file_path):

        """
        excel에 있는 데이터 읽어서 linechart 생성
        """

        wb = load_workbook(excel_file_path, data_only=True)

        for sheet_name in sheet_name_list:

            ws = wb[sheet_name]
            instance_col = [cell for cell in ws[ws.min_row] if cell.value == "INSTANCE_NUMBER"]
            print(ws, instance_col)

            for instance_num_cell in instance_col:
                row_idx = instance_num_cell.row
                col_idx = instance_num_cell.column

                datetime_cell = ws.cell(row=row_idx, column=col_idx - 1, value=None)
                avg_cell = ws.cell(row=row_idx, column=col_idx + 1, value=None)
                graph_start_cell = get_column_letter(datetime_cell.column)+"1"

                data_dict = {"min_col": avg_cell.column, "max_col": avg_cell.column,
                             "min_row": ws.min_row+1, "max_row": ws.max_row}

                category_dict = {"min_col": datetime_cell.column, "max_col": datetime_cell.column,
                                 "min_row": ws.min_row + 1, "max_row": ws.max_row}

                try:
                    ExcelUtils.insert_linechart_into_excel(ws, sheet_name, data_dict, category_dict, graph_start_cell)

                except Exception as e:
                    print("linechart insert error", e)

        wb.save(excel_file_path)
        wb.close()













